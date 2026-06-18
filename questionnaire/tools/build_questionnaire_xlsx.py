#!/usr/bin/env python3
"""Generate a branded, customer-facing Excel version of the Snowflake sizing questionnaire.

Source of truth: temp/questionnaire/questionnaire.md (transcribed below).
Output:          temp/questionnaire/snowflake-sizing-questionnaire.xlsx

Re-run after editing the questionnaire to regenerate the workbook:
    python3 tools/build_questionnaire_xlsx.py
"""
from __future__ import annotations

import os

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# --------------------------------------------------------------------------- #
# Paths
# --------------------------------------------------------------------------- #
import re

HERE = os.path.dirname(os.path.abspath(__file__))
QUESTIONNAIRE_ROOT = os.path.abspath(os.path.join(HERE, ".."))
PLUGIN_ROOT = os.path.abspath(os.path.join(HERE, "..", ".."))
REPO_ROOT = os.path.abspath(os.path.join(PLUGIN_ROOT, "..", ".."))
CHANGELOG_PATH = os.path.join(PLUGIN_ROOT, "CHANGELOG.md")
LOGO_PATH = os.path.join(
    REPO_ROOT,
    "assets",
    "snowflake-branding",
    "snowflake-icons",
    "graphic_snowflake_logo_blue.png",
)
OUTPUT_PATH = os.path.join(
    QUESTIONNAIRE_ROOT, "snowflake-sizing-questionnaire.xlsx"
)

# --------------------------------------------------------------------------- #
# Brand palette (assets/snowflake-branding/snowflake.com/palette.json)
# --------------------------------------------------------------------------- #
BLUE = "29B5E8"        # primary
DARK_BLUE = "11567F"   # header
GREY = "8A999E"
ORANGE = "FF9F36"      # accent
WHITE = "FFFFFF"
ANSWER_FILL = "EAF7FC"   # very light blue - the fill-in column
ROW_TINT = "F4FBFE"      # alternating row tint
LIGHT_GREY = "F2F4F5"

FONT = "Calibri"

# --------------------------------------------------------------------------- #
# Validation answer-type tokens
# --------------------------------------------------------------------------- #
CLOUD = ["AWS", "Azure", "Google Cloud"]
EDITION = ["Standard", "Enterprise", "Business Critical", "VPS"]
YESNO = ["Yes", "No"]
COMPLEXITY = ["Simple lookups", "Medium (joins + aggregations)", "Highly complex analytical"]
INGEST_FREQ = ["Batch", "Micro-batch", "Stream"]

# answer_type values:
#   "text"            -> free text, informational prompt only
#   "num"             -> number >= 0
#   "pct"             -> number 0..100
#   ("list", [...])   -> dropdown
#   ("list+", [...])  -> dropdown but extra free text allowed (no hard error)

# (id, category, question, reason, answer_type, unit_hint)
QUESTIONS = [
    # 1 - General & Strategic
    ("1a", "General & Strategic",
     "What are the primary business objectives for this data initiative? (e.g., reduce reporting time, enable new analytics, consolidate platforms)",
     "Helps understand the project's priority and informs architectural recommendations and potential high-value use cases.",
     "text", ""),
    ("1b", "General & Strategic",
     "Are you migrating from an existing data platform (e.g., Teradata, Oracle, Hadoop, Redshift)? If so, please specify.",
     "Informs the migration strategy and complexity, and allows more accurate estimates if the current platform's resource consumption data is available.",
     "text", ""),
    ("1c", "General & Strategic",
     'Which Cloud Provider (AWS / Azure / Google Cloud) and Region do you require? (Answer as: cloud + full region, e.g. "AWS Europe (London)")',
     "Determines the per-credit, AI-credit, and storage rates and ensures data-residency requirements are met. A fundamental component of every cost calculation.",
     ("list+", CLOUD), "cloud + full region"),
    ("1d", "General & Strategic",
     "Which Snowflake Edition do you require? (Standard / Enterprise / Business Critical / VPS)",
     "Sets the credit rate and the floor of available features. If unsure, leave blank and we infer it.",
     ("list", EDITION), ""),
    ("1e", "General & Strategic",
     "What is the proposed timeline? (development start [month/date], production go-live [month/date], and any phased rollout milestones)",
     "The go-live date and ramp shape drive the per-month consumption curve. A concrete dev-start and go-live month make the Year-1 ramp accurate instead of assumed.",
     "text", "dev start, go-live, milestones"),

    # 2 - Data Sources & Ingestion
    ("2a", "Data Sources & Ingestion",
     "What are the ongoing source systems for data? (e.g., on-prem databases, SaaS apps like Salesforce, IoT event streams, logs)",
     "Identifies the variety and complexity of data to ingest, informing the required ingestion tools and patterns (ETL/ELT).",
     "text", ""),
    ("2b", "Data Sources & Ingestion",
     "What are the formats of the incoming data? (structured CSV/Parquet, semi-structured JSON/XML, unstructured PDF/images). If known, give expected compression ratio (default 3x).",
     "Influences storage requirements and the compute needed for parsing/transformation. Format drives the compression ratio that converts raw TB to billed storage TB.",
     "text", "formats; compression ratio"),
    ("2c", "Data Sources & Ingestion",
     "For each of your largest data sources, what is the ingestion volume and frequency? ([GB/day] or [GB/month] + batch / micro-batch / stream)",
     "Directly impacts ingestion compute sizing. Volume drives ingestion credits and Snowpipe charges; frequency decides batch warehouse vs Snowpipe vs streaming.",
     "text", "[GB/day or GB/month] + frequency"),
    ("2d", "Data Sources & Ingestion",
     "Do any sources need to arrive within seconds or a few minutes of being generated? (live fraud alerts, real-time dashboards, IoT), if yes, expected [streaming GB/month].",
     "Determines whether Snowpipe Streaming (sub-second) is required vs standard Snowpipe. Streaming adds ingestion compute and the streaming GB/month sizes it.",
     ("list+", YESNO), "if yes: [streaming GB/month]"),

    # 3, Storage & Data Architecture
    ("3a", "Storage & Data Architecture",
     "What is the total raw source data volume [GB or TB] you plan to migrate initially? (if known, give expected compression ratio; default 3x)",
     "The primary driver of baseline storage cost. Compression reduces it, but raw size is the starting point for the estimate.",
     "text", "[GB or TB]; compression ratio"),
    ("3b", "Storage & Data Architecture",
     "How many [days] of historical data must be immediately available for query rollback (Time Travel)?",
     "Determines Time Travel and Fail-safe storage cost. Longer retention increases storage consumption and cost.",
     "num", "[days]"),
    ("3c", "Storage & Data Architecture",
     "Do you require separate, isolated environments for Development, Testing/QA, and Production? (which environments)",
     "Determines the number of databases/objects to manage and informs the compute strategy for multiple teams and release cycles.",
     "text", "which environments"),
    ("3d", "Storage & Data Architecture",
     "Do you have large volumes of historical data rarely or never queried after a certain age?, if yes: [% of total storage] and after how many [months] it becomes cold.",
     "Archive storage is significantly cheaper than standard. If cold data is a large proportion, enabling archive storage meaningfully reduces long-term TCO.",
     "text", "if yes: [% of storage] + [months]"),
    ("3e", "Storage & Data Architecture",
     "What proportion of your data changes (insert/update/delete) on a typical day? [% per day] (default 10%)",
     "Churn rate drives the Time Travel and Fail-safe overhead on top of base storage. Without it, overhead is assumed at 10% and may be materially wrong.",
     "pct", "[% per day]"),

    # 4, Workloads & Use Cases
    ("4a", "Workloads & Use Cases",
     "What are the operating hours for your environment overall? (weekdays 9am–6pm; 24x7; overnight batch only; mixed)",
     "A significant cost driver and the single most impactful missing input. Weekday-only = 22 days/month; 24x7 = 30 days/month.",
     "text", ""),
    ("4b", "Workloads & Use Cases",
     "List your primary workloads (Corporate BI, Ad-hoc Analytics, ELT/Transformation, Data Science), then answer 4c–4h for each one.",
     "Critical input for compute estimation. Separating workloads lets us right-size individual warehouses instead of one oversized warehouse.",
     "text", ""),
    ("4c", "Workloads & Use Cases",
     "[For each workload] How many users run queries concurrently during peak hours? [# concurrent users]",
     "Determines warehouse size and multi-cluster configuration needed to meet concurrency without queueing",
     "text", "[# concurrent users]"),
    ("4d", "Workloads & Use Cases",
     '[For each workload] What is the query-performance SLA? ([seconds] target, e.g. "dashboards < 5s", "ETL < 1hr")',
     "Drives warehouse size. Faster SLAs require larger warehouses.",
     "text", "[seconds] target"),
    ("4e", "Workloads & Use Cases",
     "[For each workload] What are the operating hours / frequency? (24x7 loading, 9am–5pm Mon–Fri BI, once daily 2am batch)",
     "Directly impacts credit consumption. Warehouses suspend when idle, so the schedule (hours/day × days/month) is key to cost.",
     "text", ""),
    ("4f", "Workloads & Use Cases",
     "[For each workload] How would you classify query complexity? (simple lookups / medium / highly complex analytical)",
     "Informs warehouse size. More complex queries need more compute (larger warehouses) to finish within the SLA.",
     ("list+", COMPLEXITY), "per workload"),
    ("4h", "Workloads & Use Cases",
     "[For each workload] Roughly how many queries per user per day [#], and typical query runtime [seconds] if known?",
     "Completes the BI/analytics credit formula (users × queries/day × runtime × size-credits × days). Without it, per-user consumption is assumed from complexity alone.",
     "text", "[# queries/user/day] + [seconds]"),

    # 5, Python, ML & Containers
    ("5a", "Python, Machine Learning & Containers",
     "Will users write and run Python directly inside the platform? (notebooks, Pandas/NumPy, scikit-learn, PySpark-style transforms)",
     "Python notebooks on standard warehouses suffice for exploration. Heavy in-memory ML benefits from Snowpark-Optimized Warehouses. Determines standard vs Snowpark-Optimized sizing.",
     ("list+", YESNO), ""),
    ("5b", "Python, Machine Learning & Containers",
     "Will you train ML models where large datasets load entirely into memory?, if yes: dataset size [GB] and training frequency.",
     "ML training at scale needs Snowpark-Optimized Warehouses. Dataset size and frequency prevent under-sizing.",
     "text", "if yes: [GB] + frequency"),
    ("5c", "Python, Machine Learning & Containers",
     "Do you need to run containerized apps/services inside the platform? (real-time inference API, web service, Docker app), if yes: app type, [requests/day], GPU or CPU.",
     "Containerized services run on Snowpark Container Services compute pools, separate from warehouses. App type, request volume, and GPU-vs-CPU are required to size the pool.",
     "text", "if yes: app type, [requests/day], GPU/CPU"),

    # 6, AI & Analytics Intelligence
    ("6a", "AI & Analytics Intelligence",
     "Would business users benefit from asking questions about data in plain English? (Cortex Analyst), if yes: [# analyst users] and [questions per user per day].",
     "Determines Cortex Analyst usage. User count × daily question frequency is the sizing input.",
     "text", "if yes: [# users] + [questions/user/day]"),
    ("6b", "AI & Analytics Intelligence",
     "Will you use an LLM to process or generate text in pipelines/apps? (summarising feedback, generating descriptions, classifying emails), if yes: [records/month] and avg text length.",
     "Determines Cortex Complete token volume. Cost = tokens × model rate.",
     "text", "if yes: [records/month] + text length"),
    ("6c", "AI & Analytics Intelligence",
     "Do you need to search across large collections of free-text documents by meaning? (contracts, tickets, KB articles), if yes: [indexed data GB] and refresh frequency.",
     "Determines Cortex Search, billed at X credits per GB of indexed data per month. The indexed GB is the key sizing input.",
     "text", "if yes: [indexed GB] + refresh freq"),
    ("6d", "AI & Analytics Intelligence",
     "Do you need to extract structured information from documents, PDFs, or images?, if yes: [documents/month] and [avg pages per document].",
     "Determines AI_EXTRACT token volume, billed per token from the AI credit pool. Document count × page length is required to estimate it.",
     "text", "if yes: [docs/month] + [pages/doc]"),
    ("6e", "AI & Analytics Intelligence",
     "Do you need text analysis on columns of data? (sentiment, classification, translation, summarisation), if yes: which functions, [# rows], avg text length.",
     "Determines which AI SQL functions are used (ai_sentiment / ai_classify / ai_translate / ai_summarize), each billed per million tokens.",
     "text", "if yes: functions, [# rows], text length"),
    ("6f", "AI & Analytics Intelligence",
     "Do you plan a conversational AI assistant / chatbot that answers by querying your data? (Cortex Agents / Snowflake Intelligence), if yes: [# users], [sessions/user/day], [messages/session].",
     "Determines Cortex Agents usage, billed per token. Usage is modeled as users × sessions/day × messages/session.",
     "text", "if yes: [# users], [sessions/day], [msgs/session]"),
    ("6g", "AI & Analytics Intelligence",
     "Do developers write SQL/Python regularly, and would an in-tool AI assistant be valuable? (Cortex Code), if yes: [# developers], surface(s) CLI / Snowsight / Desktop, [queries/dev/day].",
     "Determines Cortex Code, billed per token. Developer count, surface, and daily query volume are the key inputs.",
     "text", "if yes: [# devs], surface, [queries/dev/day]"),

    # 7, Data Governance & Security
    ("7a", "Data Governance & Security",
     "Do you have specific regulatory/compliance requirements? (HIPAA, PCI-DSS, GDPR)",
     "Determines the required Edition. Business Critical is necessary for the highest compliance levels (HIPAA, PCI).",
     "text", ""),
    ("7b", "Data Governance & Security",
     "Do you require dynamic data masking, row-level access policies, or tokenization?",
     "These are Enterprise-Edition-or-higher features. The answer sets the minimum required edition and its cost.",
     ("list+", YESNO), ""),
    ("7c", "Data Governance & Security",
     "Must all connections travel over a private internal network, not the public internet? (AWS PrivateLink, Azure Private Link), if yes, estimated [GB processed/month].",
     "PrivateLink adds a per-endpoint monthly charge plus a per-GB data-processed charge, and requires Enterprise Edition or higher.",
     ("list+", YESNO), "if yes: [GB processed/month]"),

    # 8, Disaster Recovery & Business Continuity
    ("8a", "Disaster Recovery & Business Continuity",
     "Do you require a geographically separate backup/failover copy in case the primary region is unavailable? (yes / no)",
     "Cross-region replication is a frequently overlooked cost (seed, ongoing replication compute ~4 cr/TB, egress, replica storage)",
     ("list", YESNO), ""),
    ("8b", "Disaster Recovery & Business Continuity",
     "[If yes] How often should the backup sync? And what is your production dataset at go-live [TB] and daily change volume [GB/day]?",
     "Sync frequency sets the compute cadence; dataset size drives the one-time seed transfer; daily change drives ongoing monthly compute and egress.",
     "text", "sync freq, [TB], [GB/day]"),
    ("8c", "Disaster Recovery & Business Continuity",
     '[If yes] Which regions are involved? (source → target, e.g. "AWS US East → AWS EU West")',
     "Egress varies by region pair. The pair also determines whether an Egress Cost Optimizer cache applies.",
     "text", "source → target"),

    # 9, Real-time Database Replication (CDC)
    ("9a", "Real-time Database Replication (CDC)",
     "Do you need to continuously capture and replicate row-level changes from operational databases in near real-time? (yes / no)",
     "CDC is a separate cost category: a dedicated runtime service plus a warehouse for MERGE operations.",
     ("list", YESNO), ""),
    ("9b", "Real-time Database Replication (CDC)",
     "[If yes] Which source database types? How many separate database servers? Same physical server? For Oracle: processor type (x86 / SPARC / POWER) and [# cores].",
     "Each distinct source server needs a dedicated runtime node. Oracle adds per-core licensing; the processor type sets the licensing factor (x86 = 0.5, SPARC/POWER = 1.0).",
     "text", "DB types, # servers, Oracle cores"),
    ("9c", "Real-time Database Replication (CDC)",
     "[If yes] Roughly how many [# tables] need replicating across all sources, and approximate [row changes per day]?",
     "Runtime nodes have a 600-table limit. Daily event volume sets runtime size: < 100k events/day → Medium; > 100k → Large.",
     "text", "[# tables] + [row changes/day]"),
    ("9d", "Real-time Database Replication (CDC)",
     "[If yes] Is there an existing warehouse that can be reused for the CDC MERGE step, or must a new one be provisioned? (reuse existing / provision new)",
     "The MERGE warehouse is typically 60–70% of total CDC cost. Reusing an existing warehouse avoids duplicating that cost in the estimate.",
     ("list+", ["Reuse existing", "Provision new"]), ""),

    # 10, Data Sharing & External Collaboration
    ("10a", "Data Sharing & External Collaboration",
     "Do you need to share live data with external partners, customers, or vendors outside your organisation?, if yes: [# external parties] and how often they query.",
     "External sharing to parties without their own account needs a Reader Account (billed to you). A dedicated warehouse must be sized from external user count and query frequency.",
     "text", "if yes: [# parties] + query freq"),
    ("10b", "Data Sharing & External Collaboration",
     "Are you planning to sell data products/apps via a marketplace, or build a data-powered app your customers will use?",
     "Determines whether Native Apps / Marketplace costs apply; these also require Enterprise Edition or higher.",
     "text", ""),

    # 11, Future Growth & Platform
    ("11a", "Future Growth & Platform",
     "Projected annual growth in data volume over the next 3 years? [% per year]",
     "Essential for 3-year TCO. Storage is a recurring cost that scales with data growth.",
     "pct", "[% per year]"),
    ("11b", "Future Growth & Platform",
     "Projected annual growth in number of users and/or new workloads over the next 3 years? [% per year]",
     "Essential for modeling rising compute consumption across the X-year term, ensuring the contract value is accurate for the full duration.",
     "pct", "[% per year]"),
]

# Appendix A - internal field mapping (id, SIZING_SPEC field(s), notes)
MAPPING = [
    ("1a", "assumptions[], narrative", "Strategic context; not a numeric input."),
    ("1b", "meta narrative, replication.* (if migration)", "Migration may activate replication block."),
    ("1c", "meta.cloud, meta.region", "Drives meta.credit_rate, ai_credit_rate, storage_rate_per_tb via derive-rates.py."),
    ("1d", "meta.edition", "Can be inferred from 7a/7b/7c/10b if blank."),
    ("1e", "meta.default_dev_start_month, meta.default_go_live_month, meta.default_ramp_curve; per-row workloads[].* ", "Drives the Birdbox per-month ramp."),
    ("2a", "source narrative; serverless.*, openflow.* selection", "Identifies ingestion pattern."),
    ("2b", "storage.standard.compression_ratio", "Format → compression default."),
    ("2c", "serverless.snowpipe.gb_per_month; ingestion workloads[]", "Per-source volume + frequency."),
    ("2d", "serverless.snowpipe_streaming.enabled, .uncompressed_gb_per_month", "Real-time → streaming."),
    ("3a", "storage.standard.raw_tb_year1, .compression_ratio", "Baseline storage."),
    ("3b", "storage.standard.time_travel_days", ""),
    ("3c", "workloads[] (dev/test workload rows)", "Dev environment sizing."),
    ("3d", "Archive storage assumptions / confirm_required[]", "Proportion + cold threshold."),
    ("3e", "storage.standard.churn_rate_pct", "Time-travel + fail-safe overhead."),
    ("4a", "workloads[].days_per_month (22 vs 30)", "Global operating days."),
    ("4b", "workloads[] (one row per workload)", "Workload decomposition."),
    ("4c", "workloads[].clusters_min, .clusters_max", "MCW concurrency."),
    ("4d", "workloads[].size", "SLA → size."),
    ("4e", "workloads[].hours_per_day, .days_per_month, .auto_suspend_seconds", "Per-workload schedule."),
    ("4f", "workloads[].size", "Complexity → size."),
    ("4h", "workloads[].hours_per_day (BI formula inputs)", "Queries/day × runtime."),
    ("5a", "workloads[].warehouse_type (Standard vs Snowpark-Optimized)", ""),
    ("5b", "workloads[].warehouse_type = Snowpark-Optimized, size, memory config", "Dataset GB + frequency."),
    ("5c", "spcs.* (instance type, generation, node count, hours)", "App type + requests/day + GPU/CPU."),
    ("6a", "ai_cortex.cortex_analyst.enabled, .monthly_messages", "users × questions/day × 22."),
    ("6b", "ai_cortex.cortex_complete.enabled, .model, .monthly_input_tokens_M, .monthly_output_tokens_M", "records/mo × text length → tokens."),
    ("6c", "ai_cortex.cortex_search.enabled, .indexed_data_gb", ""),
    ("6d", "ai_cortex.cortex_functions.ai_extract.enabled, .tokens_M_monthly", "docs/mo × pages → tokens."),
    ("6e", "ai_cortex.cortex_functions.{ai_sentiment,ai_classify,ai_translate,ai_summarize}.{enabled,tokens_M_monthly}", "rows × text length → tokens."),
    ("6f", "ai_cortex.cortex_agents.{enabled,monthly_users,sessions_per_user_per_day,messages_per_session,...}", "Usage-model helper fields."),
    ("6g", "ai_cortex.cortex_code.{enabled,developers,queries_per_dev_per_day,avg_tokens_per_query}", "developers × surface × queries/day."),
    ("7a", "meta.edition (→ Business Critical)", "Compliance floor."),
    ("7b", "meta.edition (→ Enterprise)", "Masking/RLS floor."),
    ("7c", "meta.edition (→ Enterprise); PrivateLink cost line + confirm_required[]", "Per-endpoint + per-GB."),
    ("8a", "replication.* enable", "DR activation."),
    ("8b", "replication.initial_TB, .monthly_change_TB, sync frequency", "Seed + ongoing."),
    ("8c", "replication.source_region, .target_region", "Egress matrix lookup."),
    ("9a", "openflow.* CDC connector enable", "CDC activation."),
    ("9b", "openflow runtime node count; Oracle licensing (cores × processor factor)", "servers + same-server? + Oracle processor type."),
    ("9c", "openflow runtime size (Medium/Large), node count (600-table limit)", "tables + events/day."),
    ("9d", "CDC MERGE workloads[] row (new) or reuse note / confirm_required[]", "Avoids 60–70% of CDC cost if reused."),
    ("10a", "collaboration.reader_accounts.{enabled,warehouse_size,hours_per_day,days_per_month}", "external parties + query frequency."),
    ("10b", "collaboration.{native_apps,marketplace}.{enabled,monthly_subscription}; meta.edition (→ Enterprise)", ""),
    ("11a", "meta.annual_growth_rate; storage.standard.annual_growth_pct", "Data growth %/yr."),
    ("11b", "meta.annual_growth_rate; per-row workloads[].growth_rate", "User/workload growth %/yr."),
]

# --------------------------------------------------------------------------- #
# Style helpers
# --------------------------------------------------------------------------- #
def fill(hex_):
    return PatternFill("solid", fgColor=hex_)


def plugin_version():
    """Read the latest version number from CHANGELOG.md (first '## [x.y.z]')."""
    try:
        with open(CHANGELOG_PATH, encoding="utf-8") as fh:
            for line in fh:
                m = re.match(r"##\s*\[([^\]]+)\]", line.strip())
                if m:
                    return m.group(1)
    except OSError:
        pass
    return "unknown"


THIN = Side(style="thin", color="D5DDE0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

WRAP_TOP = Alignment(wrap_text=True, vertical="top")
WRAP_TOP_CENTER = Alignment(wrap_text=True, vertical="top", horizontal="center")


def build():
    wb = Workbook()

    # ----------------------------------------------------------------- Cover
    cover = wb.active
    cover.title = "Welcome"
    cover.sheet_view.showGridLines = False
    cover.sheet_properties.tabColor = BLUE
    for col, width in (("A", 3), ("B", 170), ("C", 3)):
        cover.column_dimensions[col].width = width

    if os.path.exists(LOGO_PATH):
        img = XLImage(LOGO_PATH)
        # scale to a reasonable header height
        ratio = 90 / img.height if img.height else 1
        img.height = int(img.height * ratio)
        img.width = int(img.width * ratio)
        cover.add_image(img, "B2")

    cover["B8"] = "Snowflake Solution Sizing Assessment"
    cover["B8"].font = Font(name=FONT, size=24, bold=True, color=DARK_BLUE)

    cover["B10"] = (
        "The details you provide enable us to build an accurate, right-sized "
        "solution and a predictable cost estimate tailored to your goals."
    )
    cover["B10"].font = Font(name=FONT, size=14, color="333333")
    cover["B10"].alignment = Alignment(wrap_text=True, vertical="top")
    cover.row_dimensions[10].height = 45

    cover["B12"] = "How to answer"
    cover["B12"].font = Font(name=FONT, size=14, bold=True, color=BLUE)

    guidance = (
        "•  Where a question shows a unit in square brackets (e.g. [# concurrent users], "
        "[GB/day], [indexed GB]), please give a number in that unit. A rough number is far better than none.\n"
        "•  Leave an answer blank if genuinely unknown, anything blank is sized with a documented assumption and flagged for confirmation, \n"
        "•  Enter your answers in the highlighted Answer column on the 'Questionnaire' tab. Some cells offer a drop-down list."
    )
    cover["B14"] = guidance
    cover["B14"].font = Font(name=FONT, size=14, color="333333")
    cover["B14"].alignment = Alignment(wrap_text=True, vertical="top")
    cover.row_dimensions[14].height = 130

    cover["B18"] = "Tip: blue-shaded cells are for you to fill in."
    cover["B18"].font = Font(name=FONT, size=14, italic=True, color=DARK_BLUE)

    cover["B20"] = f"version {plugin_version()}"
    cover["B20"].font = Font(name=FONT, size=14, color=GREY)

    # --------------------------------------------------------- Questionnaire
    ws = wb.create_sheet("Questionnaire")
    ws.sheet_properties.tabColor = DARK_BLUE
    ws.sheet_view.showGridLines = False
    headers = ["ID", "Category", "Question", "Answer", "Comments", "Reason for Impact"]
    widths = [6, 22, 70, 38, 38, 60]
    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
        c = ws.cell(row=1, column=i, value=h)
        c.fill = fill(DARK_BLUE)
        c.font = Font(name=FONT, size=14, bold=True, color=WHITE)
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = BORDER
    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "A2"
    # "Reason for Impact" is internal context; hide it from the customer view.
    ws.column_dimensions["F"].hidden = True

    row = 2
    current_cat = None
    band = False  # alternating tint within a category
    dv_objs = []

    for (qid, cat, question, reason, atype, hint) in QUESTIONS:
        # category section header row
        if cat != current_cat:
            current_cat = cat
            band = False
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            hc = ws.cell(row=row, column=1, value=cat)
            hc.fill = fill(BLUE)
            hc.font = Font(name=FONT, size=14, bold=True, color=WHITE)
            hc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            ws.row_dimensions[row].height = 22
            for col in range(1, 7):
                ws.cell(row=row, column=col).border = BORDER
            row += 1

        tint = ROW_TINT if band else WHITE
        band = not band

        c_id = ws.cell(row=row, column=1, value=qid)
        c_cat = ws.cell(row=row, column=2, value=cat)
        c_q = ws.cell(row=row, column=3, value=question)
        c_a = ws.cell(row=row, column=4, value=None)
        c_com = ws.cell(row=row, column=5, value=None)
        c_r = ws.cell(row=row, column=6, value=reason)

        c_id.font = Font(name=FONT, size=14, bold=True, color=DARK_BLUE)
        c_id.alignment = WRAP_TOP_CENTER
        c_cat.font = Font(name=FONT, size=14, color=GREY)
        c_cat.alignment = WRAP_TOP
        c_q.font = Font(name=FONT, size=14, color="222222")
        c_q.alignment = WRAP_TOP
        c_a.font = Font(name=FONT, size=14, color="111111")
        c_a.alignment = WRAP_TOP
        c_a.fill = fill(ANSWER_FILL)
        c_com.font = Font(name=FONT, size=14, color="111111")
        c_com.alignment = WRAP_TOP
        c_com.fill = fill(ANSWER_FILL)
        c_r.font = Font(name=FONT, size=14, color="555555")
        c_r.alignment = WRAP_TOP

        for col, cell in ((1, c_id), (2, c_cat), (3, c_q), (6, c_r)):
            cell.fill = fill(tint)
            cell.border = BORDER
        c_a.border = BORDER
        c_com.border = BORDER

        # estimate row height from question length
        qlen = max(len(question), len(reason))
        ws.row_dimensions[row].height = max(38, min(150, 22 + (qlen // 55) * 20))

        # data validation
        dv = _make_validation(atype, hint)
        if dv is not None:
            ws.add_data_validation(dv)
            dv.add(c_a)

        row += 1

    ws.print_title_rows = "1:1"

    # --------------------------------------------------- Field Mapping (hidden)
    mp = wb.create_sheet("Field Mapping - internal")
    mp.sheet_state = "hidden"
    mp.sheet_view.showGridLines = False
    mp_headers = ["ID", "SIZING_SPEC field(s)", "Notes"]
    mp_widths = [8, 70, 55]
    mp["A1"] = "INTERNAL USE ONLY - questionnaire → SIZING_SPEC field mapping (Appendix A)"
    mp.merge_cells("A1:C1")
    mp["A1"].font = Font(name=FONT, size=14, bold=True, color=ORANGE)
    for i, (h, w) in enumerate(zip(mp_headers, mp_widths), start=1):
        mp.column_dimensions[get_column_letter(i)].width = w
        c = mp.cell(row=2, column=i, value=h)
        c.fill = fill(DARK_BLUE)
        c.font = Font(name=FONT, size=14, bold=True, color=WHITE)
        c.border = BORDER
    r = 3
    for (mid, field, note) in MAPPING:
        mp.cell(row=r, column=1, value=mid).font = Font(name=FONT, bold=True, color=DARK_BLUE)
        fc = mp.cell(row=r, column=2, value=field)
        nc = mp.cell(row=r, column=3, value=note)
        for col in (1, 2, 3):
            cell = mp.cell(row=r, column=col)
            cell.alignment = WRAP_TOP
            cell.border = BORDER
            if col != 1:
                cell.font = Font(name=FONT, size=14, color="333333")
        r += 1
    mp.freeze_panes = "A3"

    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    wb.save(OUTPUT_PATH)
    return OUTPUT_PATH


def _make_validation(atype, hint):
    """Return a configured DataValidation, or None for plain free text."""
    if isinstance(atype, tuple):
        kind, options = atype
        formula = '"' + ",".join(options) + '"'
        allow_blank = True
        # "list" = strict; "list+" = allow other text (no hard stop)
        dv = DataValidation(
            type="list",
            formula1=formula,
            allow_blank=allow_blank,
            showDropDown=False,
        )
        dv.showInputMessage = True
        dv.promptTitle = "Select an option"
        dv.prompt = ("Choose from the list" + (f" ; {hint}" if hint else "")
                     + (". Other values are allowed." if kind == "list+" else "."))
        if kind == "list+":
            dv.showErrorMessage = False
        else:
            dv.showErrorMessage = True
            dv.errorTitle = "Invalid entry"
            dv.error = "Please choose one of: " + ", ".join(options)
        return dv

    if atype == "num":
        dv = DataValidation(type="decimal", operator="greaterThanOrEqual",
                            formula1="0", allow_blank=True)
        dv.showInputMessage = True
        dv.promptTitle = "Enter a number"
        dv.prompt = f"Numeric value{(' - ' + hint) if hint else ''}. Leave blank if unknown."
        dv.showErrorMessage = True
        dv.errorTitle = "Numbers only"
        dv.error = "Please enter a number of 0 or greater (or leave blank)."
        return dv

    if atype == "pct":
        dv = DataValidation(type="decimal", operator="between",
                            formula1="0", formula2="100", allow_blank=True)
        dv.showInputMessage = True
        dv.promptTitle = "Enter a percentage"
        dv.prompt = f"A percentage 0–100{(' - ' + hint) if hint else ''}. Leave blank if unknown."
        dv.showErrorMessage = True
        dv.errorTitle = "0–100 only"
        dv.error = "Please enter a percentage between 0 and 100 (or leave blank)."
        return dv

    # plain text - optionally show the unit hint as guidance
    if hint:
        dv = DataValidation(type=None, allow_blank=True)
        dv.showInputMessage = True
        dv.promptTitle = "Guidance"
        dv.prompt = hint
        return dv
    return None


if __name__ == "__main__":
    path = build()
    print(f"Wrote {path}")
